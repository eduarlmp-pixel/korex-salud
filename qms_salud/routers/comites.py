from fastapi import APIRouter, Depends, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, Response, FileResponse, JSONResponse
from sqlalchemy.orm import Session
from database import get_db
from templates_config import templates
import models, auth
from datetime import datetime
import base64, os, shutil, uuid, csv, io
from xhtml2pdf import pisa
from utils.email import enviar_notificacion_acta

router = APIRouter(prefix="/comites", tags=["comites"])
UPLOAD_DIR = "static/uploads/actas"
EVIDENCIAS_DIR = "static/uploads/compromisos"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EVIDENCIAS_DIR, exist_ok=True)

TIPOS_LABEL = {
    "calidad":"Comité de Calidad","seguridad_paciente":"Seguridad del Paciente",
    "etica":"Comité de Ética","iaas":"IAAS / Infecciones",
    "farmacia":"Comité de Farmacia","historias_clinicas":"Historias Clínicas","otro":"Otro",
}
TIPOS_COLORES = {
    "calidad":"badge-blue","seguridad_paciente":"badge-red","etica":"badge-purple",
    "iaas":"badge-orange","farmacia":"badge-green","historias_clinicas":"badge-teal","otro":"badge-gray",
}

def get_logo_b64():
    if os.path.exists("static/img/logo.png"):
        with open("static/img/logo.png","rb") as f: return base64.b64encode(f.read()).decode()
    return ""

def generar_html_acta(comite, acta, asistentes, compromisos, firmas):
    logo_b64 = get_logo_b64()

    def parse_campo(texto, clave, siguiente):
        try:
            ini = texto.index(f"||{clave}||") + len(f"||{clave}||")
            fin = texto.index(f"||{siguiente}||", ini)
            return texto[ini:fin]
        except: return ""

    orden_raw = acta.orden_dia or ""
    orden_dia_limpio = orden_raw.split("||PRES||")[0] if "||PRES||" in orden_raw else orden_raw
    pres_nombre = parse_campo(orden_raw,"PRES","SEC").split("||")[0] if "||PRES||" in orden_raw else ""
    pres_cargo  = orden_raw.split("||PRES||")[1].split("||SEC||")[0].split("||")[1] if "||PRES||" in orden_raw and "||" in orden_raw.split("||PRES||")[1] else ""
    sec_nombre  = parse_campo(orden_raw,"SEC","PROC").split("||")[0] if "||SEC||" in orden_raw else ""
    sec_cargo   = orden_raw.split("||SEC||")[1].split("||PROC||")[0].split("||")[1] if "||SEC||" in orden_raw and "||" in orden_raw.split("||SEC||")[1] else ""
    proceso     = parse_campo(orden_raw,"PROC","SUA") if "||PROC||" in orden_raw else ""
    estandares  = parse_campo(orden_raw,"SUA","PER") if "||SUA||" in orden_raw else ""
    periodo_txt = parse_campo(orden_raw,"PER","HI") if "||PER||" in orden_raw else ""
    hora_ini    = parse_campo(orden_raw,"HI","HF") if "||HI||" in orden_raw else ""
    hora_fin_t  = parse_campo(orden_raw,"HF","OBJ") if "||HF||" in orden_raw else ""
    objetivo    = orden_raw.split("||OBJ||")[1] if "||OBJ||" in orden_raw else ""

    rows_asist = ""
    for i,a in enumerate(asistentes):
        rows_asist += f"<tr><td style='text-align:center;'>{i+1}</td><td>{a.nombre}</td><td>{a.cargo}</td><td style='text-align:center;'>{'X' if a.firma else ''}</td><td style='text-align:center;'>{'X' if not a.firma else ''}</td><td style='text-align:center;'>SI</td><td></td></tr>"

    rows_comp = ""
    for i,c in enumerate(compromisos):
        cumple_si = "X" if c.estado=="cumplido" else ""
        cumple_no = "X" if c.estado!="cumplido" else ""
        rows_comp += f"<tr><td style='text-align:center;'>{i+1}</td><td>{c.descripcion}</td><td>{c.responsable or '—'}</td><td style='text-align:center;'>{cumple_si}</td><td style='text-align:center;'>{cumple_no}</td><td>{c.observacion or ''}</td></tr>"

    firmas_html = ""
    for f in firmas:
        firmas_html += f'''<div style="display:inline-block;width:200px;text-align:center;margin:10px 16px;vertical-align:top;">
          <div style="border:1px solid #ddd;border-radius:4px;padding:6px;min-height:60px;">
            {'<img src="' + f.firma_imagen + '" style="max-height:60px;max-width:180px;">' if f.firma_imagen else '<div style="height:50px;"></div>'}
          </div>
          <div style="font-size:10px;margin-top:4px;font-weight:bold;">{f.nombre}</div>
          <div style="font-size:9px;color:#555;">{f.cargo}</div>
          <div style="font-size:9px;color:#1a56a0;font-weight:bold;">{f.rol_en_acta.upper()}</div>
          {'<div style="font-size:9px;color:#b91c1c;">NO VIABLE: ' + f.comentario + '</div>' if not f.valida else ''}
        </div>'''

    num_acta = acta.numero_acta if acta else "—"
    fecha_str = acta.fecha.strftime('%d de %B de %Y') if acta and acta.fecha else "—"
    tipo_label = TIPOS_LABEL.get(comite.tipo.value,"—") if comite else "—"
    desarrollo_txt = acta.desarrollo or ""

    html = f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<style>
  body{{font-family:Arial,sans-serif;font-size:10px;margin:0;padding:20px;color:#111;}}
  .header-table{{width:100%;border-collapse:collapse;margin-bottom:0;}}
  .header-table td{{border:1px solid #888;padding:5px 8px;}}
  .header-top{{display:flex;justify-content:space-between;align-items:stretch;border:1px solid #888;margin-bottom:0;}}
  .logo-cell{{padding:6px 10px;border-right:1px solid #888;display:flex;align-items:center;}}
  .title-cell{{flex:1;text-align:center;padding:6px;border-right:1px solid #888;}}
  .code-cell{{padding:6px 10px;font-size:9px;}}
  h2{{margin:0;font-size:12px;font-weight:bold;text-transform:uppercase;}}
  h3{{margin:2px 0;font-size:11px;font-weight:bold;text-transform:uppercase;}}
  table{{width:100%;border-collapse:collapse;margin-bottom:0;}}
  th{{background:#d0dff0;padding:5px 8px;font-size:9px;text-align:left;border:1px solid #888;}}
  td{{padding:4px 8px;border:1px solid #888;font-size:10px;vertical-align:top;}}
  .sec-title{{background:#d0dff0;font-weight:bold;font-size:10px;padding:4px 8px;border:1px solid #888;text-align:center;}}
  .footer{{margin-top:10px;font-size:8px;color:#aaa;text-align:center;}}
</style></head><body>

<!-- ENCABEZADO formato GER-MCC-FM-01 -->
<div class="header-top">
  <div class="logo-cell">{'<img src="data:image/png;base64,' + logo_b64 + '" style="height:45px;">' if logo_b64 else '<div style="width:60px;height:45px;background:#1a56a0;border-radius:4px;"></div>'}</div>
  <div class="title-cell">
    <h3>{tipo_label.upper()}</h3>
    <h2>FORMATO ACTA DE COMITÉ</h2>
  </div>
  <div class="code-cell">
    <div><b>Código:</b> GER-MCC-FM-01</div>
    <div><b>Versión:</b> 01</div>
    <div><b>Fecha:</b> {datetime.utcnow().strftime('%b. %Y').upper()}</div>
  </div>
</div>

<!-- 1. INFORMACIÓN GENERAL -->
<table style="margin-top:0;">
  <tr><td colspan="6" class="sec-title">1. INFORMACIÓN GENERAL DEL COMITÉ</td></tr>
  <tr><td colspan="6"><b>{comite.nombre.upper() if comite else ''}</b></td></tr>
  <tr>
    <td style="width:120px;"><b>Proceso que lidera</b></td><td>{proceso}</td>
    <td style="width:130px;"><b>Estándares SUA relacionados</b></td><td>{estandares}</td>
    <td style="width:100px;"><b>N° Acta</b></td><td>{num_acta}</td>
  </tr>
  <tr>
    <td><b>Período</b></td><td>{periodo_txt}</td>
    <td><b>Fecha ejecución</b></td><td>{fecha_str}</td>
    <td><b>Lugar</b></td><td>{acta.lugar if acta else '—'}</td>
  </tr>
  <tr>
    <td><b>Hora inicio</b></td><td>{hora_ini}</td>
    <td><b>Hora fin</b></td><td>{hora_fin_t}</td>
    <td><b>Quórum</b></td><td>{'SÍ ✓' if acta and acta.quorum else 'NO'}</td>
  </tr>
  <tr><td colspan="2" class="sec-title">2. PRESIDENTE</td><td colspan="4"></td></tr>
  <tr><td><b>Nombre</b></td><td>{pres_nombre}</td><td><b>Cargo</b></td><td colspan="3">{pres_cargo}</td></tr>
  <tr><td colspan="2" class="sec-title">3. SECRETARIO</td><td colspan="4"></td></tr>
  <tr><td><b>Nombre</b></td><td>{sec_nombre}</td><td><b>Cargo</b></td><td colspan="3">{sec_cargo}</td></tr>
</table>

<!-- 4. CONTROL DE ASISTENCIA -->
<table style="margin-top:4px;">
  <tr><td colspan="7" class="sec-title">4. CONTROL DE ASISTENCIA</td></tr>
  <tr>
    <th style="width:30px;">N°</th><th>Nombre integrante</th><th>Cargo / Actividad</th>
    <th style="width:60px;text-align:center;">Asistencial</th>
    <th style="width:60px;text-align:center;">Administrativo</th>
    <th style="width:50px;text-align:center;">Asistió</th>
    <th style="width:80px;text-align:center;">Firma</th>
  </tr>
  {rows_asist if rows_asist else '<tr><td colspan="7" style="text-align:center;color:#888;">Sin asistentes registrados</td></tr>'}
</table>

<!-- 5. SEGUIMIENTO COMPROMISOS -->
<table style="margin-top:4px;">
  <tr><td colspan="6" class="sec-title">5. SEGUIMIENTO A COMPROMISOS DEL COMITÉ ANTERIOR</td></tr>
  <tr>
    <th style="width:30px;">No</th><th>Compromiso</th><th>Responsable</th>
    <th style="width:40px;text-align:center;">Cumple SI</th>
    <th style="width:40px;text-align:center;">Cumple NO</th>
    <th>Observaciones</th>
  </tr>
  {rows_comp if rows_comp else '<tr><td colspan="6" style="text-align:center;color:#888;">Sin compromisos anteriores</td></tr>'}
</table>

<!-- 6. OBJETIVO -->
<table style="margin-top:4px;">
  <tr><td class="sec-title">6. OBJETIVOS DEL COMITÉ</td></tr>
  <tr><td style="min-height:40px;">{objetivo or '—'}</td></tr>
</table>

<!-- 7. DESARROLLO -->
<table style="margin-top:4px;">
  <tr><td class="sec-title">7. DESARROLLO DEL COMITÉ</td></tr>
  <tr><td style="min-height:60px;white-space:pre-wrap;">{desarrollo_txt or '—'}</td></tr>
</table>

<!-- INDICADORES -->
<table style="margin-top:4px;">
  <tr><td colspan="4" class="sec-title">8. INDICADORES DE EFECTIVIDAD DEL COMITÉ</td></tr>
  <tr><th>No</th><th>Indicador</th><th style="width:80px;text-align:center;">Resultado</th><th style="width:80px;text-align:center;">Meta</th></tr>
  <tr><td>1</td><td>Porcentaje de asistencia: N° asistentes / N° integrantes según resolución</td>
    <td style="text-align:center;">{round(len([a for a in asistentes])/max(len(asistentes),1)*100,1) if asistentes else '—'}%</td>
    <td style="text-align:center;">100%</td></tr>
  <tr><td>2</td><td>Seguimiento a compromisos: N° compromisos cumplidos / total compromisos</td>
    <td style="text-align:center;">{round(len([c for c in compromisos if c.estado=='cumplido'])/max(len(compromisos),1)*100,1) if compromisos else '—'}%</td>
    <td style="text-align:center;">100%</td></tr>
</table>

<!-- PRÓXIMA REUNIÓN -->
<table style="margin-top:4px;">
  <tr><td style="width:200px;font-weight:bold;background:#d0dff0;border:1px solid #888;">CONVOCATORIA PRÓXIMA REUNIÓN / FECHA:</td>
  <td style="border:1px solid #888;">___________________________________</td></tr>
  <tr><td style="font-weight:bold;background:#d0dff0;border:1px solid #888;">ACTA ELABORADA POR:</td>
  <td style="border:1px solid #888;">{sec_nombre.upper()} – {sec_cargo.upper()}</td></tr>
</table>

<!-- FIRMAS -->
{"<div style='margin-top:16px;'><div class='sec-title' style='margin-bottom:10px;'>FIRMAS DE VALIDACIÓN</div>" + firmas_html + "</div>" if firmas else ""}

<!-- ARCHIVO ADJUNTO -->
{f"<table style='margin-top:8px;'><tr><td class='sec-title'>ARCHIVO ADJUNTO</td></tr><tr><td style='font-size:10px;'>📎 <b>{acta.archivo_nombre}</b> — Este documento tiene un archivo adjunto. Descárgalo por separado desde la plataforma QMS Salud.</td></tr></table>" if acta and acta.archivo_nombre else ""}

<div class="footer">Acta N° {num_acta} · {tipo_label} · Generado por QMS Salud · {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}</div>
</body></html>'''
    return html

@router.get("", response_class=HTMLResponse)
async def lista_comites(request: Request, db: Session = Depends(get_db),
                         current_user=Depends(auth.get_current_user)):
    if not current_user: return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "comites", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    comites = db.query(models.Comite).filter_by(activo=True).all()
    comites_data = []
    for c in comites:
        total_actas = len(c.actas)
        ultima_acta = sorted(c.actas, key=lambda a: a.fecha, reverse=True)[0] if c.actas else None
        compromisos_pendientes = db.query(models.CompromisoComite).join(models.ActaComite).filter(
            models.ActaComite.comite_id == c.id, models.CompromisoComite.estado == "pendiente").count()
        comites_data.append({"comite":c,"total_actas":total_actas,"ultima_acta":ultima_acta,
            "compromisos_pendientes":compromisos_pendientes,
            "color":TIPOS_COLORES.get(c.tipo.value,"badge-gray"),
            "label":TIPOS_LABEL.get(c.tipo.value,c.tipo.value)})
    return templates.TemplateResponse("comites/lista.html",{
        "request":request,"user":current_user,"comites_data":comites_data,"tipos_label":TIPOS_LABEL,
        "permisos":auth.get_permisos(db,current_user)})

@router.get("/nuevo-comite", response_class=HTMLResponse)
async def nuevo_comite_form(request: Request, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user: return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "comites", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    return templates.TemplateResponse("comites/form_comite.html",{"request":request,"user":current_user,
        "permisos":auth.get_permisos(db,current_user)})

@router.post("/nuevo-comite")
async def crear_comite(request: Request, nombre: str = Form(...), tipo: str = Form(...),
    descripcion: str = Form(""), periodicidad: str = Form("mensual"),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    c = models.Comite(nombre=nombre,tipo=tipo,descripcion=descripcion,periodicidad=periodicidad)
    db.add(c); db.commit()
    return RedirectResponse(url="/comites", status_code=302)

@router.get("/{comite_id}", response_class=HTMLResponse)
async def detalle_comite(comite_id: int, request: Request, db: Session = Depends(get_db),
                          current_user=Depends(auth.get_current_user)):
    if not current_user: return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "comites", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    comite = db.query(models.Comite).filter_by(id=comite_id).first()
    if not comite: return RedirectResponse(url="/comites")
    actas = db.query(models.ActaComite).filter_by(comite_id=comite_id).order_by(models.ActaComite.fecha.desc()).all()
    compromisos_pendientes = db.query(models.CompromisoComite).join(models.ActaComite).filter(
        models.ActaComite.comite_id==comite_id, models.CompromisoComite.estado=="pendiente").all()
    return templates.TemplateResponse("comites/detalle.html",{
        "request":request,"user":current_user,"comite":comite,"actas":actas,
        "compromisos_pendientes":compromisos_pendientes,
        "label":TIPOS_LABEL.get(comite.tipo.value,comite.tipo.value),
        "color":TIPOS_COLORES.get(comite.tipo.value,"badge-gray"),
        "permisos":auth.get_permisos(db,current_user)})

@router.get("/{comite_id}/nueva-acta", response_class=HTMLResponse)
async def nueva_acta_form(comite_id: int, request: Request, db: Session = Depends(get_db),
                           current_user=Depends(auth.get_current_user)):
    if not current_user: return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "comites", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    comite = db.query(models.Comite).filter_by(id=comite_id).first()
    empleados = db.query(models.Empleado).filter_by(activo=True).all()
    return templates.TemplateResponse("comites/form_acta.html",{
        "request":request,"user":current_user,"comite":comite,"empleados":empleados,
        "label":TIPOS_LABEL.get(comite.tipo.value,comite.tipo.value),
        "permisos":auth.get_permisos(db,current_user)})

@router.post("/{comite_id}/nueva-acta")
async def crear_acta(comite_id: int, request: Request,
    numero_acta: str = Form(...), fecha: str = Form(...),
    lugar: str = Form(""), orden_dia: str = Form(""), desarrollo: str = Form(""),
    estado: str = Form("realizado"), quorum: str = Form("off"),
    presidente_nombre: str = Form(""), presidente_cargo: str = Form(""),
    secretario_nombre: str = Form(""), secretario_cargo: str = Form(""),
    proceso_lidera: str = Form(""), estandares_sua: str = Form(""),
    periodo: str = Form(""), hora_inicio: str = Form(""), hora_fin: str = Form(""),
    objetivo: str = Form(""),
    archivo: UploadFile = File(None),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):

    archivo_nombre=None; archivo_ruta=None
    if archivo and archivo.filename:
        archivo_nombre=f"{comite_id}_{numero_acta}_{archivo.filename}".replace(" ","_")
        archivo_ruta=f"{UPLOAD_DIR}/{archivo_nombre}"
        with open(archivo_ruta,"wb") as f: shutil.copyfileobj(archivo.file,f)

    # Guardar datos extra en orden_dia con separador
    extra = f"||PRES||{presidente_nombre}||{presidente_cargo}||SEC||{secretario_nombre}||{secretario_cargo}||PROC||{proceso_lidera}||SUA||{estandares_sua}||PER||{periodo}||HI||{hora_inicio}||HF||{hora_fin}||OBJ||{objetivo}"
    orden_completo = (orden_dia or "") + extra

    acta = models.ActaComite(
        comite_id=comite_id, numero_acta=numero_acta,
        fecha=datetime.fromisoformat(fecha), lugar=lugar,
        orden_dia=orden_completo, desarrollo=desarrollo,
        estado=estado, quorum=(quorum=="on"),
        archivo_nombre=archivo_nombre, archivo_ruta=archivo_ruta,
        creado_por=current_user.id)
    db.add(acta); db.commit(); db.refresh(acta)

    # Copiar asistentes de la última acta
    ultima_acta = db.query(models.ActaComite).filter(
        models.ActaComite.comite_id == comite_id,
        models.ActaComite.id != acta.id
    ).order_by(models.ActaComite.fecha.desc()).first()

    if ultima_acta and ultima_acta.asistentes:
        import uuid
        for asis in ultima_acta.asistentes:
            nuevo_asis = models.AsistenteComite(
                acta_id=acta.id,
                nombre=asis.nombre,
                cargo=asis.cargo,
                area=asis.area,
                cedula=asis.cedula,
                email=asis.email,
                empleado_id=asis.empleado_id,
                rol_en_acta=asis.rol_en_acta,
                token_firma=str(uuid.uuid4())
            )
            db.add(nuevo_asis)
        db.commit()

    return RedirectResponse(url=f"/comites/{comite_id}", status_code=302)

@router.get("/{comite_id}/acta/{acta_id}", response_class=HTMLResponse)
async def detalle_acta(comite_id: int, acta_id: int, request: Request,
                        db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user: return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "comites", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    acta = db.query(models.ActaComite).filter_by(id=acta_id).first()
    comite = db.query(models.Comite).filter_by(id=comite_id).first()
    firmas = db.query(models.FirmaActa).filter_by(acta_id=acta_id).all()
    acta_bloqueada = any(f.rol_en_acta in ("presidente","secretario") for f in firmas) and \
                     sum(1 for f in firmas if f.rol_en_acta in ("presidente","secretario")) >= 2
    return templates.TemplateResponse("comites/acta.html",{
        "request":request,"user":current_user,"acta":acta,"comite":comite,
        "firmas":firmas,"acta_bloqueada":acta_bloqueada,
        "label":TIPOS_LABEL.get(comite.tipo.value,comite.tipo.value),
        "color":TIPOS_COLORES.get(comite.tipo.value,"badge-gray"),
        "permisos":auth.get_permisos(db,current_user)})

@router.get("/api/empleado/{cedula}")
async def buscar_empleado(cedula: str, db: Session = Depends(get_db)):
    emp = db.query(models.Empleado).filter_by(documento=cedula, activo=True).first()
    if not emp:
        return JSONResponse({"found": False})
    return JSONResponse({
        "found": True,
        "id": emp.id,
        "nombres": emp.nombres,
        "apellidos": emp.apellidos,
        "nombre_completo": f"{emp.nombres} {emp.apellidos}",
        "cargo": emp.cargo or "",
        "area": emp.area or "",
        "email": emp.email or "",
    })


@router.post("/{comite_id}/acta/{acta_id}/asistente")
async def agregar_asistente(comite_id: int, acta_id: int,
    cedula: str = Form(""), nombre: str = Form(...), cargo: str = Form(""),
    area: str = Form(""), email: str = Form(""), empleado_id: str = Form(""),
    rol_en_acta: str = Form("integrante"),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):

    if not current_user:
        return RedirectResponse(url="/login")

    # Evitar duplicados por cédula en el mismo acta
    if cedula:
        existente = db.query(models.AsistenteComite).filter_by(acta_id=acta_id, cedula=cedula).first()
        if existente:
            return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)

    token = str(uuid.uuid4())
    emp_id = int(empleado_id) if empleado_id.isdigit() else None

    asistente = models.AsistenteComite(
        acta_id=acta_id, nombre=nombre, cargo=cargo, area=area,
        cedula=cedula, email=email, empleado_id=emp_id,
        rol_en_acta=rol_en_acta, token_firma=token,
    )
    db.add(asistente)
    db.commit()

    if email:
        acta = db.query(models.ActaComite).filter_by(id=acta_id).first()
        comite = db.query(models.Comite).filter_by(id=comite_id).first()
        if acta and comite:
            asistentes = db.query(models.AsistenteComite).filter_by(acta_id=acta_id).all()
            compromisos = db.query(models.CompromisoComite).filter_by(acta_id=acta_id).all()
            firmas = db.query(models.FirmaActa).filter_by(acta_id=acta_id).all()
            
            html_acta = generar_html_acta(comite, acta, asistentes, compromisos, firmas)
            pdf_buf = io.BytesIO()
            pisa.CreatePDF(io.StringIO(html_acta), dest=pdf_buf)
            pdf_bytes = pdf_buf.getvalue()

            enviar_notificacion_acta(
                nombre=nombre, email_destino=email,
                numero_acta=acta.numero_acta,
                nombre_comite=comite.nombre,
                token=token, comite_id=comite_id, acta_id=acta_id,
                pdf_bytes=pdf_bytes, pdf_name=f"Acta_{acta.numero_acta}.pdf"
            )

    return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)


@router.post("/{comite_id}/acta/{acta_id}/asistente/{asistente_id}/enviar-correo")
async def enviar_correo_asistente(comite_id: int, acta_id: int, asistente_id: int,
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    
    if not current_user:
        return RedirectResponse(url="/login")
        
    asistente = db.query(models.AsistenteComite).filter_by(id=asistente_id, acta_id=acta_id).first()
    if not asistente or not asistente.email:
        return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)
        
    acta = db.query(models.ActaComite).filter_by(id=acta_id).first()
    comite = db.query(models.Comite).filter_by(id=comite_id).first()
    
    if acta and comite:
        asistentes = db.query(models.AsistenteComite).filter_by(acta_id=acta_id).all()
        compromisos = db.query(models.CompromisoComite).filter_by(acta_id=acta_id).all()
        firmas = db.query(models.FirmaActa).filter_by(acta_id=acta_id).all()
        
        pdf_bytes = None
        try:
            from routers.comites import generar_html_acta
            import xhtml2pdf.pisa as pisa
            from io import BytesIO
            html_pdf = generar_html_acta(comite, acta, asistentes, compromisos, firmas)
            result = BytesIO()
            pisa_status = pisa.CreatePDF(BytesIO(html_pdf.encode("utf-8")), dest=result)
            if not pisa_status.err:
                pdf_bytes = result.getvalue()
        except Exception as e:
            print(f"Error generando PDF para adjunto manual: {e}")
            
        from utils.email import enviar_notificacion_acta
        pdf_name = f"Acta_{acta.numero_acta}_{comite.nombre.replace(' ','_')}.pdf"
        enviar_notificacion_acta(
            nombre=asistente.nombre,
            email_destino=asistente.email,
            numero_acta=acta.numero_acta,
            nombre_comite=comite.nombre,
            token=asistente.token_firma,
            comite_id=comite_id,
            acta_id=acta_id,
            pdf_bytes=pdf_bytes,
            pdf_name=pdf_name
        )
        
    return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)


@router.get("/firmar", response_class=HTMLResponse)
async def firmar_por_token(request: Request, token: str, db: Session = Depends(get_db)):
    asistente = db.query(models.AsistenteComite).filter_by(token_firma=token).first()
    if not asistente:
        return HTMLResponse("<h2>Enlace inválido o expirado.</h2>", status_code=404)
    if asistente.firma:
        return HTMLResponse("<h2>Ya has firmado este acta. Gracias.</h2>")
    acta = db.query(models.ActaComite).filter_by(id=asistente.acta_id).first()
    comite = db.query(models.Comite).filter_by(id=acta.comite_id).first() if acta else None
    firmas_existentes = db.query(models.FirmaActa).filter_by(acta_id=asistente.acta_id).all()
    pres_sec = sum(1 for f in firmas_existentes if f.rol_en_acta in ("presidente", "secretario"))
    bloqueada = pres_sec >= 2
    return templates.TemplateResponse("comites/firmar.html", {
        "request": request, "asistente": asistente,
        "acta": acta, "comite": comite,
        "bloqueada": bloqueada,
        "label": TIPOS_LABEL.get(comite.tipo.value, comite.tipo.value) if comite else "",
    })


@router.post("/firmar")
async def guardar_firma_token(request: Request,
    token: str = Form(...), firma_imagen: str = Form(""),
    valida: str = Form("true"), comentario: str = Form(""),
    db: Session = Depends(get_db)):
    asistente = db.query(models.AsistenteComite).filter_by(token_firma=token).first()
    if not asistente or asistente.firma:
        return HTMLResponse("<h2>Enlace inválido, expirado o ya utilizado.</h2>", status_code=400)
    acta = db.query(models.ActaComite).filter_by(id=asistente.acta_id).first()
    if not acta:
        return HTMLResponse("<h2>Acta no encontrada.</h2>", status_code=404)

    firmas_existentes = db.query(models.FirmaActa).filter_by(acta_id=acta.id).all()
    pres_sec = sum(1 for f in firmas_existentes if f.rol_en_acta in ("presidente", "secretario"))
    if pres_sec >= 2:
        return HTMLResponse("<h2>El acta ya está bloqueada. No se pueden agregar más firmas.</h2>")

    db.add(models.FirmaActa(
        acta_id=acta.id,
        nombre=asistente.nombre, cargo=asistente.cargo or "",
        rol_en_acta=asistente.rol_en_acta,
        firma_imagen=firma_imagen, valida=(valida == "true"),
        comentario=comentario,
    ))
    asistente.firma = True
    db.commit()
    return templates.TemplateResponse("comites/firma_ok.html", {
        "request": request, "nombre": asistente.nombre,
    })


@router.post("/{comite_id}/acta/{acta_id}/firma")
async def agregar_firma(comite_id: int, acta_id: int,
    asistente_id: int = Form(...),
    valida: str = Form("true"), comentario: str = Form(""),
    firma_imagen: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):

    if not current_user:
        return RedirectResponse(url="/login")

    asistente = db.query(models.AsistenteComite).filter_by(id=asistente_id, acta_id=acta_id).first()
    if not asistente:
        return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)
    if asistente.firma:
        return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)

    firmas = db.query(models.FirmaActa).filter_by(acta_id=acta_id).all()
    pres_sec = sum(1 for f in firmas if f.rol_en_acta in ("presidente", "secretario"))
    if pres_sec >= 2:
        return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)

    db.add(models.FirmaActa(
        acta_id=acta_id,
        nombre=asistente.nombre, cargo=asistente.cargo or "",
        rol_en_acta=asistente.rol_en_acta,
        firma_imagen=firma_imagen, valida=(valida == "true"),
        comentario=comentario,
    ))
    asistente.firma = True
    db.commit()
    return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)

@router.get("/plantilla-integrantes")
async def plantilla_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Integrantes"
    cols = [
        ("A", "Cédula (Documento)",      20),
        ("B", "Nombres",                 18),
        ("C", "Apellidos",               18),
        ("D", "Cargo",                   20),
        ("E", "Área",                    16),
        ("F", "Correo electrónico",      28),
        ("G", "Rol en acta",             18),
    ]
    fill = PatternFill("solid", fgColor="1E3A5F")
    font_h = Font(bold=True, color="FFFFFF")
    for col, label, width in cols:
        cell = ws[f"{col}1"]
        cell.value = label
        cell.font = font_h
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[col].width = width
    ws.append(["1098765432", "Juan", "Pérez", "Médico", "Urgencias", "juan@hospital.com", "integrante"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_integrantes.xlsx"})


@router.post("/{comite_id}/acta/{acta_id}/importar-integrantes")
async def importar_integrantes(comite_id: int, acta_id: int,
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")

    contenido = await archivo.read()
    ext = archivo.filename.rsplit(".", 1)[-1].lower() if archivo.filename else ""
    filas = []
    if ext in ("xls", "xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contenido))
            ws = wb.active
            cabecera = [str(c.value or "").strip().lower().replace("é","e").replace("á","a")
                        .replace("ó","o").replace("í","i").replace("ú","u").replace(" ","_")
                        .replace("(documento)","").replace("(cedula)","").strip()
                        for c in ws[1]]
            for fila in ws.iter_rows(min_row=2, values_only=True):
                filas.append({cabecera[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(fila)})
        except Exception:
            return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)
    else:
        texto = contenido.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(texto))
        filas = [{k.strip().lower(): v.strip() for k, v in fila.items()} for fila in reader]

    acta = db.query(models.ActaComite).filter_by(id=acta_id).first()
    comite = db.query(models.Comite).filter_by(id=comite_id).first()

    agregados = 0
    for fila in filas:
        cedula = (fila.get("cedula") or fila.get("cedula_") or "").strip()
        if not cedula:
            continue
        # Evitar duplicados
        if db.query(models.AsistenteComite).filter_by(acta_id=acta_id, cedula=cedula).first():
            continue

        # Buscar empleado en BD (fuente de verdad: tabla empleados)
        emp = db.query(models.Empleado).filter_by(documento=cedula, activo=True).first()

        # Soporta columnas separadas (nombres + apellidos) o columna única (nombre)
        nombres_col   = (fila.get("nombres") or "").strip()
        apellidos_col = (fila.get("apellidos") or "").strip()
        nombre = (fila.get("nombre") or "").strip()
        if nombres_col or apellidos_col:
            nombre = f"{nombres_col} {apellidos_col}".strip()
        cargo  = (fila.get("cargo")  or "").strip()
        area   = (fila.get("area")   or fila.get("area_") or "").strip()
        email  = (fila.get("correo_electronico") or fila.get("email") or "").strip()
        rol    = (fila.get("rol_en_acta") or fila.get("rol") or "integrante").strip()

        if emp:
            nombre = nombre or f"{emp.nombres} {emp.apellidos}"
            cargo  = cargo  or (emp.cargo or "")
            area   = area   or (emp.area  or "")
            email  = email  or (emp.email or "")

        if not nombre:
            continue

        token = str(uuid.uuid4())
        db.add(models.AsistenteComite(
            acta_id=acta_id, cedula=cedula, nombre=nombre,
            cargo=cargo, area=area, email=email,
            empleado_id=emp.id if emp else None,
            rol_en_acta=rol, token_firma=token,
        ))
        agregados += 1

        if email and acta and comite:
            asistentes = db.query(models.AsistenteComite).filter_by(acta_id=acta_id).all()
            compromisos = db.query(models.CompromisoComite).filter_by(acta_id=acta_id).all()
            firmas = db.query(models.FirmaActa).filter_by(acta_id=acta_id).all()
            
            html_acta = generar_html_acta(comite, acta, asistentes, compromisos, firmas)
            pdf_buf = io.BytesIO()
            pisa.CreatePDF(io.StringIO(html_acta), dest=pdf_buf)
            pdf_bytes = pdf_buf.getvalue()

            enviar_notificacion_acta(
                nombre=nombre, email_destino=email,
                numero_acta=acta.numero_acta,
                nombre_comite=comite.nombre,
                token=token, comite_id=comite_id, acta_id=acta_id,
                pdf_bytes=pdf_bytes, pdf_name=f"Acta_{acta.numero_acta}.pdf"
            )

    db.commit()
    return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)


@router.post("/{comite_id}/acta/{acta_id}/compromiso")
async def agregar_compromiso(comite_id: int, acta_id: int,
    descripcion: str = Form(...), responsable: str = Form(""), fecha_limite: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    db.add(models.CompromisoComite(acta_id=acta_id, descripcion=descripcion,
        responsable=responsable,
        fecha_limite=datetime.fromisoformat(fecha_limite) if fecha_limite else None))
    db.commit()
    return RedirectResponse(url=f"/comites/{comite_id}/acta/{acta_id}", status_code=302)

@router.post("/compromiso/{comp_id}/estado")
async def cambiar_estado_compromiso(comp_id: int, estado: str = Form(...),
    observacion: str = Form(""), db: Session = Depends(get_db),
    current_user=Depends(auth.get_current_user)):
    comp = db.query(models.CompromisoComite).filter_by(id=comp_id).first()
    if comp:
        comp.estado=estado; comp.observacion=observacion; db.commit()
        acta = db.query(models.ActaComite).filter_by(id=comp.acta_id).first()
        return RedirectResponse(url=f"/comites/{acta.comite_id}/acta/{acta.id}", status_code=302)
    return RedirectResponse(url="/comites", status_code=302)


@router.post("/compromiso/{comp_id}/evidencia")
async def subir_evidencia_compromiso(comp_id: int,
    evidencia: UploadFile = File(...),
    observacion: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    comp = db.query(models.CompromisoComite).filter_by(id=comp_id).first()
    if not comp:
        return RedirectResponse(url="/comites", status_code=302)

    if evidencia and evidencia.filename:
        ext = os.path.splitext(evidencia.filename)[1].lower()
        nombre_unico = f"comp_{comp_id}_{uuid.uuid4().hex}{ext}"
        ruta = os.path.join(EVIDENCIAS_DIR, nombre_unico)
        with open(ruta, "wb") as f:
            shutil.copyfileobj(evidencia.file, f)
        comp.evidencia_ruta = ruta.replace("\\", "/")
        comp.evidencia_nombre = evidencia.filename

    comp.estado = "cumplido"
    comp.fecha_cumplimiento = datetime.utcnow()
    if observacion:
        comp.observacion = observacion
    db.commit()

    acta = db.query(models.ActaComite).filter_by(id=comp.acta_id).first()
    return RedirectResponse(url=f"/comites/{acta.comite_id}/acta/{acta.id}", status_code=302)


@router.get("/compromiso/{comp_id}/evidencia/descargar")
async def descargar_evidencia(comp_id: int, db: Session = Depends(get_db),
                               current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    comp = db.query(models.CompromisoComite).filter_by(id=comp_id).first()
    if not comp or not comp.evidencia_ruta or not os.path.exists(comp.evidencia_ruta):
        raise HTTPException(status_code=404, detail="Evidencia no encontrada")
    return FileResponse(comp.evidencia_ruta,
        filename=comp.evidencia_nombre or os.path.basename(comp.evidencia_ruta),
        media_type="application/octet-stream")

@router.get("/acta/{acta_id}/descargar")
async def descargar_acta(acta_id: int, db: Session = Depends(get_db),
                          current_user=Depends(auth.get_current_user)):
    acta = db.query(models.ActaComite).filter_by(id=acta_id).first()
    if acta and acta.archivo_ruta and os.path.exists(acta.archivo_ruta):
        return FileResponse(acta.archivo_ruta, filename=acta.archivo_nombre)
    return RedirectResponse(url="/comites")

@router.get("/{comite_id}/acta/{acta_id}/pdf")
async def pdf_acta(comite_id: int, acta_id: int, db: Session = Depends(get_db),
                    current_user=Depends(auth.get_current_user)):
    acta = db.query(models.ActaComite).filter_by(id=acta_id).first()
    comite = db.query(models.Comite).filter_by(id=comite_id).first()
    asistentes = db.query(models.AsistenteComite).filter_by(acta_id=acta_id).all()
    compromisos = db.query(models.CompromisoComite).filter_by(acta_id=acta_id).all()
    firmas = db.query(models.FirmaActa).filter_by(acta_id=acta_id).all()

    html = generar_html_acta(comite, acta, asistentes, compromisos, firmas)

    num_acta = acta.numero_acta if acta else "—"
    return Response(content=html, media_type="text/html",
        headers={"Content-Disposition": f"attachment; filename=acta_{num_acta}_{comite_id}.html"})
