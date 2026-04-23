from fastapi import APIRouter, Depends, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from templates_config import templates
from sqlalchemy.orm import Session
from database import get_db
import models, auth
from datetime import datetime
import csv, io

# ─── RIESGOS ──────────────────────────────────────────────────────────────────
riesgos_router = APIRouter(prefix="/riesgos", tags=["riesgos"])

NIVEL_MAP = {
    (1,1):"bajo",(1,2):"bajo",(2,1):"bajo",(2,2):"bajo",
    (2,3):"moderado",(3,2):"moderado",(3,3):"moderado",
    (3,4):"alto",(4,3):"alto",(4,4):"alto",(2,4):"alto",(4,2):"moderado",
    (4,5):"critico",(5,4):"critico",(5,5):"critico",(3,5):"critico",(5,3):"alto",(1,5):"alto",(5,1):"alto"
}

@riesgos_router.get("", response_class=HTMLResponse)
async def lista_riesgos(request: Request, nivel: str = "", db: Session = Depends(get_db),
                         current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "riesgos", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    query = db.query(models.Riesgo).filter_by(activo=True)
    if nivel:
        query = query.filter(models.Riesgo.nivel == nivel)
    riesgos = query.order_by(models.Riesgo.probabilidad.desc(), models.Riesgo.impacto.desc()).all()
    # Mapa térmico 5x5
    mapa = {(p, i): [] for p in range(1, 6) for i in range(1, 6)}
    for r in db.query(models.Riesgo).filter_by(activo=True).all():
        if r.probabilidad and r.impacto:
            mapa[(r.probabilidad, r.impacto)].append(r)
    return templates.TemplateResponse("riesgos/lista.html", {
        "request": request, "user": current_user,
        "riesgos": riesgos, "mapa": mapa, "filtro_nivel": nivel,
        "permisos": auth.get_permisos(db, current_user),
    })

@riesgos_router.get("/nuevo", response_class=HTMLResponse)
async def nuevo_riesgo_form(request: Request, db: Session = Depends(get_db),
                             current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "riesgos", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    usuarios = db.query(models.Usuario).filter_by(activo=True).all()
    return templates.TemplateResponse("riesgos/form.html", {
        "request": request, "user": current_user, "usuarios": usuarios, "riesgo": None,
        "permisos": auth.get_permisos(db, current_user),
    })

@riesgos_router.post("/nuevo")
async def crear_riesgo(
    request: Request,
    codigo: str = Form(...), proceso: str = Form(...),
    descripcion: str = Form(...), causa: str = Form(""),
    consecuencia: str = Form(""), probabilidad: int = Form(...),
    impacto: int = Form(...), control_existente: str = Form(""),
    accion_propuesta: str = Form(""), responsable_id: int = Form(...),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    nivel = NIVEL_MAP.get((probabilidad, impacto), "moderado")
    r = models.Riesgo(
        codigo=codigo, proceso=proceso, descripcion=descripcion,
        causa=causa, consecuencia=consecuencia,
        probabilidad=probabilidad, impacto=impacto, nivel=nivel,
        control_existente=control_existente, accion_propuesta=accion_propuesta,
        responsable_id=responsable_id
    )
    db.add(r)
    db.commit()
    return RedirectResponse(url="/riesgos", status_code=302)

# ─── INDICADORES ──────────────────────────────────────────────────────────────
indicadores_router = APIRouter(prefix="/indicadores", tags=["indicadores"])

@indicadores_router.get("", response_class=HTMLResponse)
async def lista_indicadores(request: Request, db: Session = Depends(get_db),
                             current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "indicadores", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    indicadores = db.query(models.Indicador).filter_by(activo=True).all()
    inds_data = []
    for ind in indicadores:
        mediciones = db.query(models.MedicionIndicador).filter_by(
            indicador_id=ind.id).order_by(models.MedicionIndicador.periodo).limit(6).all()
        ultima = mediciones[-1] if mediciones else None
        estado = "sin_datos"
        if ultima:
            if ind.meta_minima is not None and ind.meta_maxima is not None:
                estado = "cumple" if ind.meta_minima <= ultima.valor <= ind.meta_maxima else "alerta"
        inds_data.append({
            "indicador": ind, "mediciones": mediciones,
            "ultima": ultima, "estado": estado
        })
    return templates.TemplateResponse("indicadores/lista.html", {
        "request": request, "user": current_user, "indicadores": inds_data,
        "permisos": auth.get_permisos(db, current_user),
    })

@indicadores_router.get("/nuevo", response_class=HTMLResponse)
async def nuevo_indicador_form(request: Request, db: Session = Depends(get_db),
                                current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "indicadores", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    usuarios = db.query(models.Usuario).filter_by(activo=True).all()
    return templates.TemplateResponse("indicadores/form.html", {
        "request": request, "user": current_user, "usuarios": usuarios, "indicador": None,
        "permisos": auth.get_permisos(db, current_user),
    })

@indicadores_router.post("/nuevo")
async def crear_indicador(
    request: Request,
    codigo: str = Form(...), nombre: str = Form(...),
    proceso: str = Form(""), formula: str = Form(""),
    unidad: str = Form(""), meta: float = Form(0),
    meta_minima: float = Form(0), meta_maxima: float = Form(100),
    frecuencia: str = Form("mensual"), responsable_id: int = Form(...),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    ind = models.Indicador(
        codigo=codigo, nombre=nombre, proceso=proceso, formula=formula,
        unidad=unidad, meta=meta, meta_minima=meta_minima, meta_maxima=meta_maxima,
        frecuencia=frecuencia, responsable_id=responsable_id
    )
    db.add(ind)
    db.commit()
    return RedirectResponse(url="/indicadores", status_code=302)

@indicadores_router.post("/{ind_id}/medicion")
async def registrar_medicion(
    ind_id: int, valor: float = Form(...),
    periodo: str = Form(...), observacion: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    m = models.MedicionIndicador(
        indicador_id=ind_id, valor=valor, periodo=periodo,
        observacion=observacion, registrado_por=current_user.id
    )
    db.add(m)
    db.commit()
    return RedirectResponse(url="/indicadores", status_code=302)

# ─── TALENTO HUMANO ───────────────────────────────────────────────────────────
talento_router = APIRouter(prefix="/talento", tags=["talento"])

@talento_router.get("", response_class=HTMLResponse)
async def lista_empleados(request: Request, q: str = "", area: str = "",
                           db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    query = db.query(models.Empleado).filter_by(activo=True)
    if q:
        query = query.filter(
            models.Empleado.nombres.ilike(f"%{q}%") |
            models.Empleado.apellidos.ilike(f"%{q}%") |
            models.Empleado.cargo.ilike(f"%{q}%")
        )
    if area:
        query = query.filter(models.Empleado.area == area)
    empleados = query.order_by(models.Empleado.apellidos).all()
    areas = db.query(models.Empleado.area).distinct().all()

    from datetime import timedelta
    hoy = datetime.utcnow()
    en30 = hoy + timedelta(days=30)
    vencidos = db.query(models.VencimientoCurso).join(models.Empleado).filter(
        models.VencimientoCurso.activo == True,
        models.Empleado.activo == True,
        models.VencimientoCurso.fecha_vencimiento < hoy,
    ).count()
    proximos = db.query(models.VencimientoCurso).join(models.Empleado).filter(
        models.VencimientoCurso.activo == True,
        models.Empleado.activo == True,
        models.VencimientoCurso.fecha_vencimiento >= hoy,
        models.VencimientoCurso.fecha_vencimiento <= en30,
    ).count()
    caps_proximas = db.query(models.CapacitacionPlan).filter(
        models.CapacitacionPlan.activo == True,
        models.CapacitacionPlan.estado == "programada",
    ).order_by(models.CapacitacionPlan.fecha_programada).limit(5).all()

    return templates.TemplateResponse("talento/lista.html", {
        "request": request, "user": current_user,
        "empleados": empleados, "areas": [a[0] for a in areas if a[0]],
        "filtros": {"q": q, "area": area},
        "vencidos": vencidos, "proximos": proximos,
        "caps_proximas": caps_proximas, "hoy": hoy,
        "permisos": auth.get_permisos(db, current_user),
    })

@talento_router.get("/nuevo", response_class=HTMLResponse)
async def nuevo_empleado_form(request: Request, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    return templates.TemplateResponse("talento/form.html", {
        "request": request, "user": current_user, "empleado": None,
        "permisos": auth.get_permisos(db, current_user),
    })

@talento_router.post("/nuevo")
async def crear_empleado(
    request: Request,
    nombres: str = Form(...), apellidos: str = Form(...),
    documento: str = Form(...), cargo: str = Form(""),
    area: str = Form(""), tipo_contrato: str = Form(""),
    fecha_ingreso: str = Form(""), rethus: str = Form(""),
    email: str = Form(""), telefono: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    emp = models.Empleado(
        nombres=nombres, apellidos=apellidos, documento=documento,
        cargo=cargo, area=area, tipo_contrato=tipo_contrato,
        rethus=rethus, email=email, telefono=telefono,
        fecha_ingreso=datetime.fromisoformat(fecha_ingreso) if fecha_ingreso else None
    )
    db.add(emp)
    db.commit()
    return RedirectResponse(url="/talento", status_code=302)

@talento_router.get("/plantilla-capacitaciones")
async def plantilla_capacitaciones(current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capacitaciones"

    secciones = [
        ("INFORMACIÓN DE LA CAPACITACIÓN", "1A5276", [
            ("Título *",                          32),
            ("Tipo",                              18),
            ("Descripción / Objetivo",            38),
            ("Capacitador / Instructor *",        28),
            ("Cédula del Capacitador",            22),
            ("Institución / Entidad",             26),
        ]),
        ("LOGÍSTICA", "1E8449", [
            ("Fecha (AAAA-MM-DD) *",              22),
            ("Hora inicio (HH:MM)",               16),
            ("Duración (horas)",                  16),
            ("Modalidad",                         18),
            ("Lugar / Plataforma",                26),
        ]),
        ("DIRIGIDA A — se busca en base de datos de Talento", "6C3483", [
            ("Área / Servicio  (ej: Urgencias)",  28),
            ("Cargo / Profesión  (opcional)",     28),
        ]),
        ("RESPONSABLE / SEGUIMIENTO", "784212", [
            ("Cédula del Responsable Interno",    26),
            ("Estado (programada / realizada)",   24),
        ]),
    ]

    col_idx = 1
    for nombre_sec, color, columnas in secciones:
        start = col_idx
        end   = col_idx + len(columnas) - 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        sc = ws.cell(row=1, column=start)
        sc.value = nombre_sec
        sc.font  = Font(bold=True, color="FFFFFF", size=11)
        sc.fill  = PatternFill("solid", fgColor=color)
        sc.alignment = Alignment(horizontal="center", vertical="center")
        for etiqueta, ancho in columnas:
            hc = ws.cell(row=2, column=col_idx)
            hc.value = etiqueta
            hc.font  = Font(bold=True, color="FFFFFF", size=9)
            hc.fill  = PatternFill("solid", fgColor="2D6A9F")
            hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho
            col_idx += 1

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    # Validación desplegable para Tipo y Modalidad
    dv_tipo = DataValidation(type="list",
        formula1='"SST,Calidad,Obligatoria,Técnica,Humanización,Emergencias,Farmacología,Otra"',
        allow_blank=True)
    dv_mod = DataValidation(type="list",
        formula1='"presencial,virtual,mixta"', allow_blank=True)
    dv_est = DataValidation(type="list",
        formula1='"programada,realizada"', allow_blank=True)
    ws.add_data_validation(dv_tipo)
    ws.add_data_validation(dv_mod)
    ws.add_data_validation(dv_est)
    dv_tipo.sqref = "B3:B500"
    dv_mod.sqref  = "J3:J500"
    dv_est.sqref  = "N3:N500"

    # Fila ejemplo
    ws.append([
        "BLS Soporte Vital Básico 2025",                       # título
        "Emergencias",                                          # tipo
        "Reentrenamiento anual en RCP para personal asistencial",  # descripción
        "Dr. Andrés Ramírez",                                   # capacitador
        "12345678",                                             # cédula capacitador
        "Cruz Roja Colombiana",                                 # institución
        "2025-06-15",                                           # fecha
        "08:00",                                                # hora
        "8",                                                    # horas
        "presencial",                                           # modalidad
        "Auditorio principal",                                  # lugar
        "Urgencias",                                            # área dirigida
        "Médico General",                                       # cargo (opcional)
        "",                                                     # cédula responsable
        "programada",                                           # estado
    ])

    # Hoja de referencia de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("CAMPO", "DESCRIPCIÓN", "EJEMPLO"),
        ("Título *",              "Nombre de la capacitación (obligatorio)",        "BLS Soporte Vital 2025"),
        ("Tipo",                  "SST / Calidad / Obligatoria / Técnica / Humanización / Emergencias / Farmacología / Otra", "Emergencias"),
        ("Descripción",           "Objetivo o contenido de la capacitación",        "Reentrenamiento en RCP"),
        ("Capacitador *",         "Nombre del instructor o capacitador",            "Dr. Andrés Ramírez"),
        ("Cédula del Capacitador","Cédula del capacitador — el sistema trae su cargo y área automáticamente", "12345678"),
        ("Institución",           "Entidad que imparte la capacitación",            "Cruz Roja"),
        ("Fecha *",               "Formato AAAA-MM-DD",                             "2025-06-15"),
        ("Hora inicio",           "Formato HH:MM (24h)",                            "08:00"),
        ("Duración",              "Número de horas",                                "8"),
        ("Modalidad",             "presencial / virtual / mixta",                   "presencial"),
        ("Lugar",                 "Salón, plataforma o dirección",                  "Auditorio principal"),
        ("Área / Servicio",       "El sistema busca en Talento Humano TODOS los empleados activos de esa área", "Urgencias"),
        ("Cargo / Profesión",     "Opcional: filtra dentro del área solo ese cargo. Dejar vacío = todos del área", "Médico General"),
        ("Cédula Responsable",    "Cédula del responsable interno — se trae de Talento Humano",  "87654321"),
        ("Estado",                "programada (por realizarse) / realizada (ya ocurrió)",        "programada"),
    ]
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 30
    fill_h2 = PatternFill("solid", fgColor="1A5276")
    for i, fila in enumerate(instrucciones, 1):
        for j, val in enumerate(fila, 1):
            c = ws2.cell(row=i, column=j, value=val)
            if i == 1:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = fill_h2
            c.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_capacitaciones.xlsx"})


@talento_router.post("/importar-capacitaciones")
async def importar_capacitaciones(archivo: UploadFile = File(...),
                                   db: Session = Depends(get_db),
                                   current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)

    contenido = await archivo.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        fila1 = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        header_row = 2 if any(len(v) > 4 and v[0].isdigit() is False and "información" in v.lower() or "logística" in v.lower() for v in fila1) else 1
        cabecera = [str(ws.cell(header_row, c).value or "").strip().lower()
                    for c in range(1, ws.max_column + 1)]
        filas = []
        for fila in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(v is None or str(v).strip() == "" for v in fila):
                continue
            filas.append({cabecera[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(fila)})
    except Exception:
        return RedirectResponse(url="/talento/capacitaciones?error=excel", status_code=302)

    def g(f, *keys):
        for k in keys:
            v = f.get(k, "").strip()
            if v and v.lower() not in ("none", "nan"):
                return v
        return ""

    insertadas = 0
    for f in filas:
        titulo = g(f, "título *", "titulo *", "título", "titulo")
        if not titulo:
            continue

        fecha_str = g(f, "fecha (aaaa-mm-dd) *", "fecha (aaaa-mm-dd)", "fecha")
        hora_str  = g(f, "hora inicio (hh:mm)", "hora inicio", "hora")
        fecha_dt  = _parse_fecha(fecha_str)
        if fecha_dt and hora_str:
            try:
                h, m = hora_str.split(":")
                from datetime import timedelta
                fecha_dt = fecha_dt.replace(hour=int(h), minute=int(m))
            except Exception:
                pass

        dur = g(f, "duración (horas)", "duración", "duracion")
        try:
            dur_h = float(dur) if dur else 0
        except Exception:
            dur_h = 0

        estado = g(f, "estado (programada/realizada)", "estado") or "programada"
        if estado not in ("programada", "realizada"):
            estado = "programada"

        cedula_resp = g(f, "cédula del responsable interno", "cedula del responsable interno", "cedula responsable")
        responsable = db.query(models.Empleado).filter_by(documento=cedula_resp, activo=True).first() if cedula_resp else None

        # Si el capacitador tiene cédula, completar nombre desde Talento Humano
        ced_cap = g(f, "cédula del capacitador", "cedula del capacitador")
        nombre_cap = g(f, "capacitador / instructor *", "capacitador / instructor", "capacitador")
        if ced_cap:
            emp_cap = db.query(models.Empleado).filter_by(documento=ced_cap, activo=True).first()
            if emp_cap:
                nombre_cap = f"{emp_cap.nombres} {emp_cap.apellidos}"

        cap = models.CapacitacionPlan(
            titulo=titulo,
            tipo=g(f, "tipo"),
            descripcion=g(f, "descripción / objetivo", "descripcion / objetivo", "descripcion"),
            capacitador=nombre_cap,
            institucion_cap=g(f, "institución / entidad", "institucion / entidad", "institución", "institucion"),
            fecha_programada=fecha_dt,
            duracion_horas=dur_h,
            modalidad=g(f, "modalidad") or "presencial",
            lugar=g(f, "lugar / plataforma", "lugar"),
            dirigida_a=g(f, "área completa (ej: urgencias)", "area completa"),
            responsable_id=responsable.id if responsable else None,
            estado=estado,
            creado_por=current_user.id,
        )
        db.add(cap)
        db.flush()

        # Inscribir desde base de datos de Talento Humano por área y/o cargo
        area_val  = g(f, "área / servicio  (ej: urgencias)", "área / servicio", "area / servicio")
        cargo_val = g(f, "cargo / profesión  (opcional)", "cargo / profesión", "cargo / profesion")

        if area_val or cargo_val:
            q_emps = db.query(models.Empleado).filter(models.Empleado.activo == True)
            if area_val:
                q_emps = q_emps.filter(models.Empleado.area.ilike(f"%{area_val}%"))
            if cargo_val:
                q_emps = q_emps.filter(
                    models.Empleado.cargo.ilike(f"%{cargo_val}%") |
                    models.Empleado.profesion.ilike(f"%{cargo_val}%")
                )
            for emp in q_emps.all():
                db.add(models.CapacitacionAsistente(capacitacion_id=cap.id, empleado_id=emp.id))

        insertadas += 1

    db.commit()
    return RedirectResponse(url=f"/talento/capacitaciones?importadas={insertadas}", status_code=302)


@talento_router.get("/capacitaciones", response_class=HTMLResponse)
async def lista_capacitaciones_plan(request: Request, estado: str = "",
                                     db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    q = db.query(models.CapacitacionPlan).filter_by(activo=True)
    if estado:
        q = q.filter_by(estado=estado)
    caps = q.order_by(models.CapacitacionPlan.fecha_programada.desc()).all()
    return templates.TemplateResponse("talento/capacitaciones_plan.html", {
        "request": request, "user": current_user,
        "capacitaciones": caps, "estado": estado,
        "permisos": auth.get_permisos(db, current_user),
    })

@talento_router.get("/capacitaciones/nueva", response_class=HTMLResponse)
async def nueva_capacitacion_form(request: Request, db: Session = Depends(get_db),
                                   current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    empleados = db.query(models.Empleado).filter_by(activo=True).order_by(models.Empleado.apellidos).all()
    return templates.TemplateResponse("talento/capacitacion_form.html", {
        "request": request, "user": current_user,
        "empleados": empleados, "cap": None,
        "permisos": auth.get_permisos(db, current_user),
    })

@talento_router.post("/capacitaciones/nueva")
async def crear_capacitacion(
    titulo: str = Form(...), descripcion: str = Form(""),
    tipo: str = Form(""), fecha_programada: str = Form(""),
    duracion_horas: float = Form(0), modalidad: str = Form("presencial"),
    lugar: str = Form(""), responsable_id: str = Form(""),
    capacitador: str = Form(""), institucion_cap: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    if not current_user:
        return RedirectResponse(url="/login")
    cap = models.CapacitacionPlan(
        titulo=titulo, descripcion=descripcion, tipo=tipo,
        capacitador=capacitador, institucion_cap=institucion_cap,
        duracion_horas=duracion_horas, modalidad=modalidad, lugar=lugar,
        responsable_id=int(responsable_id) if responsable_id else None,
        creado_por=current_user.id,
        fecha_programada=datetime.fromisoformat(fecha_programada) if fecha_programada else None,
    )
    db.add(cap)
    db.commit()
    return RedirectResponse(url=f"/talento/capacitaciones/{cap.id}", status_code=302)

@talento_router.get("/capacitaciones/{cap_id}", response_class=HTMLResponse)
async def detalle_capacitacion(cap_id: int, request: Request,
                                db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    cap = db.query(models.CapacitacionPlan).filter_by(id=cap_id, activo=True).first()
    if not cap:
        return RedirectResponse(url="/talento/capacitaciones", status_code=302)
    empleados_todos = db.query(models.Empleado).filter_by(activo=True).order_by(models.Empleado.apellidos).all()
    ids_inscritos = {a.empleado_id for a in cap.asistentes}
    areas = db.query(models.Empleado.area).filter_by(activo=True).distinct().all()
    return templates.TemplateResponse("talento/capacitacion_detalle.html", {
        "request": request, "user": current_user,
        "cap": cap, "empleados_todos": empleados_todos,
        "ids_inscritos": ids_inscritos,
        "areas": [a[0] for a in areas if a[0]],
        "permisos": auth.get_permisos(db, current_user),
    })

@talento_router.post("/capacitaciones/{cap_id}/inscribir")
async def inscribir_empleado(cap_id: int, empleado_id: int = Form(...),
                              db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    existe = db.query(models.CapacitacionAsistente).filter_by(capacitacion_id=cap_id, empleado_id=empleado_id).first()
    if not existe:
        db.add(models.CapacitacionAsistente(capacitacion_id=cap_id, empleado_id=empleado_id))
        db.commit()
    return RedirectResponse(url=f"/talento/capacitaciones/{cap_id}", status_code=302)

@talento_router.post("/capacitaciones/{cap_id}/inscribir-area")
async def inscribir_por_area(cap_id: int, area: str = Form(...),
                              db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    empleados = db.query(models.Empleado).filter_by(activo=True, area=area).all()
    for emp in empleados:
        existe = db.query(models.CapacitacionAsistente).filter_by(capacitacion_id=cap_id, empleado_id=emp.id).first()
        if not existe:
            db.add(models.CapacitacionAsistente(capacitacion_id=cap_id, empleado_id=emp.id))
    db.commit()
    return RedirectResponse(url=f"/talento/capacitaciones/{cap_id}", status_code=302)

@talento_router.post("/capacitaciones/{cap_id}/asistente/{ast_id}/actualizar")
async def actualizar_asistente(cap_id: int, ast_id: int,
                                asistio: str = Form("1"), aprobado: str = Form("1"),
                                nota: str = Form(""),
                                db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    ast = db.query(models.CapacitacionAsistente).filter_by(id=ast_id, capacitacion_id=cap_id).first()
    if ast:
        ast.asistio = asistio == "1"
        ast.aprobado = aprobado == "1"
        ast.nota = float(nota) if nota else None
        db.commit()
    return RedirectResponse(url=f"/talento/capacitaciones/{cap_id}", status_code=302)

@talento_router.post("/capacitaciones/{cap_id}/asistente/{ast_id}/quitar")
async def quitar_asistente(cap_id: int, ast_id: int,
                            db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    ast = db.query(models.CapacitacionAsistente).filter_by(id=ast_id, capacitacion_id=cap_id).first()
    if ast:
        db.delete(ast)
        db.commit()
    return RedirectResponse(url=f"/talento/capacitaciones/{cap_id}", status_code=302)

@talento_router.post("/capacitaciones/{cap_id}/cerrar")
async def cerrar_capacitacion(cap_id: int,
                               db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_editar"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    cap = db.query(models.CapacitacionPlan).filter_by(id=cap_id).first()
    if cap and cap.estado == "programada":
        cap.estado = "realizada"
        for ast in cap.asistentes:
            if ast.asistio:
                db.add(models.Capacitacion(
                    empleado_id=ast.empleado_id,
                    tema=cap.titulo, tipo=cap.tipo or "Institucional",
                    fecha=cap.fecha_programada, horas=cap.duracion_horas or 0,
                    institucion=cap.lugar or "Institución",
                    aprobado=ast.aprobado,
                ))
        db.commit()
    return RedirectResponse(url=f"/talento/capacitaciones/{cap_id}", status_code=302)

@talento_router.post("/capacitaciones/{cap_id}/cancelar")
async def cancelar_capacitacion(cap_id: int,
                                 db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    cap = db.query(models.CapacitacionPlan).filter_by(id=cap_id).first()
    if cap:
        cap.estado = "cancelada"
        db.commit()
    return RedirectResponse(url=f"/talento/capacitaciones/{cap_id}", status_code=302)

@talento_router.get("/vencimientos", response_class=HTMLResponse)
async def dashboard_vencimientos(request: Request, area: str = "", estado: str = "",
                                  db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    hoy = datetime.utcnow()
    from datetime import timedelta
    en30 = hoy + timedelta(days=30)
    en60 = hoy + timedelta(days=60)
    query = db.query(models.VencimientoCurso).join(models.Empleado).filter(
        models.VencimientoCurso.activo == True,
        models.Empleado.activo == True,
    )
    if area:
        query = query.filter(models.Empleado.area == area)
    vencimientos = query.order_by(models.VencimientoCurso.fecha_vencimiento).all()
    vencidos = [v for v in vencimientos if v.fecha_vencimiento < hoy]
    proximos = [v for v in vencimientos if hoy <= v.fecha_vencimiento <= en30]
    atencion = [v for v in vencimientos if en30 < v.fecha_vencimiento <= en60]
    vigentes = [v for v in vencimientos if v.fecha_vencimiento > en60]
    areas = db.query(models.Empleado.area).filter_by(activo=True).distinct().all()
    return templates.TemplateResponse("talento/vencimientos.html", {
        "request": request, "user": current_user,
        "vencidos": vencidos, "proximos": proximos,
        "atencion": atencion, "vigentes": vigentes,
        "hoy": hoy, "area": area,
        "areas": [a[0] for a in areas if a[0]],
        "permisos": auth.get_permisos(db, current_user),
    })

def _parse_fecha(val):
    if not val:
        return None
    val = str(val).strip()
    if not val or val.lower() in ("none", "nan", ""):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(val)
    except Exception:
        return None

def _agregar_vencimiento(db, emp_id, nombre, categoria, fecha_real, fecha_vec):
    if not fecha_vec:
        return
    db.add(models.VencimientoCurso(
        empleado_id=emp_id, nombre_curso=nombre, categoria=categoria,
        fecha_realizacion=fecha_real, fecha_vencimiento=fecha_vec,
    ))

@talento_router.get("/plantilla-empleados")
async def plantilla_empleados(current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personal"

    # (nombre_seccion, color_hex, [(etiqueta_columna, ancho), ...])
    secciones = [
        ("1. DATOS DE IDENTIFICACIÓN", "1E3A5F", [
            ("Cédula / ID",                    18),
            ("Apellidos",                      20),
            ("Nombres",                        20),
            ("Profesión",                      22),
            ("Registro Profesional / ReTHUS",  28),
            ("Área / Servicio",                20),
            ("Cargo",                          20),
            ("Tipo de Contrato",               20),
            ("Fecha Ingreso (AAAA-MM-DD)",     24),
            ("Correo Electrónico",             28),
            ("Teléfono",                       16),
        ]),
        ("2. FORMACIÓN ACADÉMICA", "2D6A9F", [
            ("Título Obtenido",                26),
            ("Institución Educativa",          26),
            ("Año de Graduación",              18),
        ]),
        ("3. CURSOS OBLIGATORIOS Y CERTIFICACIONES", "1A5276", [
            ("BLS - Fecha Realización",        24),
            ("BLS - Fecha Vencimiento",        24),
            ("ACLS - Fecha Realización",       24),
            ("ACLS - Fecha Vencimiento",       24),
            ("Violencia Sexual - Fecha Cert.", 28),
            ("Duelo / Humanización - Fecha",   28),
            ("Inmunobiológicos - Fecha",       24),
            ("Residuos Hospitalarios - Fecha", 28),
        ]),
        ("4. SEGURIDAD Y SALUD EN EL TRABAJO (SST)", "0E6655", [
            ("Hepatitis B - Fecha",            22),
            ("Tétano - Fecha",                 18),
            ("Influenza - Fecha",              18),
            ("COVID-19 - Fecha",               18),
            ("Talla EPP / Uniforme",           18),
            ("Examen Médico Ocupacional",      26),
        ]),
    ]

    col_idx = 1
    for nombre_sec, color, columnas in secciones:
        start = col_idx
        end = col_idx + len(columnas) - 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        sec_cell = ws.cell(row=1, column=start)
        sec_cell.value = nombre_sec
        sec_cell.font = Font(bold=True, color="FFFFFF", size=11)
        sec_cell.fill = PatternFill("solid", fgColor=color)
        sec_cell.alignment = Alignment(horizontal="center", vertical="center")
        for etiqueta, ancho in columnas:
            hcell = ws.cell(row=2, column=col_idx)
            hcell.value = etiqueta
            hcell.font = Font(bold=True, color="FFFFFF", size=9)
            hcell.fill = PatternFill("solid", fgColor="2D6A9F")
            hcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho
            col_idx += 1

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 38
    ws.freeze_panes = "A3"

    ws.append([
        "12345678", "Pérez García", "Juan Carlos", "Médico General", "RM-12345",
        "Urgencias", "Médico de Turno", "Indefinido", "2023-01-15",
        "juan@hospital.co", "3001234567",
        "Medicina General", "Universidad Nacional", "2015",
        "2024-01-15", "2026-01-15",
        "2024-03-10", "2026-03-10",
        "2023-06-01", "2023-08-15", "", "2024-02-20",
        "2020-01-01", "2021-05-10", "2024-11-01", "2023-04-15",
        "M / 42", "2024-08-30",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=plantilla_empleados.xlsx"})


@talento_router.post("/importar-empleados")
async def importar_empleados(archivo: UploadFile = File(...),
                              db: Session = Depends(get_db),
                              current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)

    contenido = await archivo.read()
    ext = (archivo.filename or "").rsplit(".", 1)[-1].lower()
    filas = []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        # Detectar fila de cabecera: la plantilla tiene sección en fila 1, cabeceras en fila 2
        fila1 = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        header_row = 2 if any("identificaci" in v.lower() or "formaci" in v.lower() for v in fila1) else 1
        cabecera = [str(ws.cell(header_row, c).value or "").strip().lower()
                    for c in range(1, ws.max_column + 1)]
        for fila in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(v is None or str(v).strip() == "" for v in fila):
                continue
            filas.append({cabecera[i]: (str(v).strip() if v is not None else "")
                          for i, v in enumerate(fila)})
    except Exception:
        return RedirectResponse(url="/talento?error=excel", status_code=302)

    def col(f, *keys):
        for k in keys:
            v = f.get(k, "").strip()
            if v:
                return v
        return ""

    insertados, actualizados = 0, 0
    from datetime import timedelta

    for f in filas:
        doc = col(f, "cédula / id", "cedula / id", "cedula", "documento", "id")
        if not doc:
            continue

        emp = db.query(models.Empleado).filter_by(documento=doc).first()
        es_nuevo = emp is None

        if es_nuevo:
            emp = models.Empleado(documento=doc)
            db.add(emp)

        emp.apellidos          = col(f, "apellidos") or (emp.apellidos if not es_nuevo else "")
        emp.nombres            = col(f, "nombres") or (emp.nombres if not es_nuevo else "")
        emp.profesion          = col(f, "profesión", "profesion") or emp.profesion
        emp.rethus             = col(f, "registro profesional / rethus", "rethus") or emp.rethus
        emp.area               = col(f, "área / servicio", "area / servicio", "area") or emp.area
        emp.cargo              = col(f, "cargo") or emp.cargo
        emp.tipo_contrato      = col(f, "tipo de contrato", "tipo_contrato") or emp.tipo_contrato
        emp.email              = col(f, "correo electrónico", "correo electronico", "email") or emp.email
        emp.telefono           = col(f, "teléfono", "telefono") or emp.telefono
        emp.titulo_obtenido    = col(f, "título obtenido", "titulo obtenido") or emp.titulo_obtenido
        emp.institucion_formacion = col(f, "institución educativa", "institucion educativa") or emp.institucion_formacion
        emp.anio_graduacion    = col(f, "año de graduación", "año de graduacion") or emp.anio_graduacion
        emp.talla_epp          = col(f, "talla epp / uniforme", "talla epp") or emp.talla_epp
        emp.fecha_ingreso      = _parse_fecha(col(f, "fecha ingreso (aaaa-mm-dd)", "fecha ingreso")) or emp.fecha_ingreso
        emp.fecha_examen_ocupacional = _parse_fecha(col(f, "examen médico ocupacional", "examen medico ocupacional")) or emp.fecha_examen_ocupacional

        if not emp.apellidos:
            emp.apellidos = "—"
        if not emp.nombres:
            emp.nombres = "—"

        db.flush()  # obtener emp.id si es nuevo

        # Cursos con vencimiento explícito (BLS, ACLS)
        bls_real = _parse_fecha(col(f, "bls - fecha realización", "bls - fecha realizacion"))
        bls_vec  = _parse_fecha(col(f, "bls - fecha vencimiento"))
        _agregar_vencimiento(db, emp.id, "BLS - Soporte Vital Básico", "BLS", bls_real, bls_vec)

        acls_real = _parse_fecha(col(f, "acls - fecha realización", "acls - fecha realizacion"))
        acls_vec  = _parse_fecha(col(f, "acls - fecha vencimiento"))
        _agregar_vencimiento(db, emp.id, "ACLS - Soporte Vital Avanzado", "ACLS", acls_real, acls_vec)

        # Cursos con fecha única → vencimiento calculado automáticamente
        _CURSOS_AUTO = [
            ("violencia sexual - fecha cert.", "Atenc. Víctimas Violencia Sexual", "Certificación", 3 * 365),
            ("duelo / humanización - fecha",   "Manejo de Duelo / Humanización",  "Certificación", 5 * 365),
            ("inmunobiológicos - fecha",        "Administración Inmunobiológicos", "SST",           2 * 365),
            ("residuos hospitalarios - fecha",  "Gestión Residuos Hospitalarios",  "SST",           365),
            ("hepatitis b - fecha",             "Vacuna Hepatitis B",              "Vacunación",    10 * 365),
            ("tétano - fecha",                  "Vacuna Tétano",                   "Vacunación",    10 * 365),
            ("influenza - fecha",               "Vacuna Influenza",                "Vacunación",    365),
            ("covid-19 - fecha",                "Vacuna COVID-19",                 "Vacunación",    365),
        ]
        for col_key, nombre_curso, categoria, dias_validad in _CURSOS_AUTO:
            fecha_r = _parse_fecha(col(f, col_key))
            if fecha_r:
                _agregar_vencimiento(db, emp.id, nombre_curso, categoria,
                                     fecha_r, fecha_r + timedelta(days=dias_validad))

        # Examen médico (anual)
        fecha_exam = emp.fecha_examen_ocupacional
        if fecha_exam:
            _agregar_vencimiento(db, emp.id, "Examen Médico Ocupacional", "SST",
                                 fecha_exam, fecha_exam + timedelta(days=365))

        if es_nuevo:
            insertados += 1
        else:
            actualizados += 1

    db.commit()
    return RedirectResponse(url=f"/talento?importados={insertados}&actualizados={actualizados}", status_code=302)


@talento_router.get("/plantilla-cursos")
async def plantilla_cursos(current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cursos"
    headers_info = [
        ("A", "Documento empleado (Cédula)", 26),
        ("B", "Nombre del curso",            30),
        ("C", "Categoría",                   22),
        ("D", "Fecha realización (AAAA-MM-DD)", 28),
        ("E", "Fecha vencimiento (AAAA-MM-DD)", 28),
        ("F", "Institución",                 24),
        ("G", "Observación",                 30),
    ]
    fill = PatternFill("solid", fgColor="1A5276")
    font_h = Font(bold=True, color="FFFFFF")
    for col, label, width in headers_info:
        cell = ws[f"{col}1"]
        cell.value = label
        cell.font = font_h
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[col].width = width
    ws.append(["12345678", "BLS Soporte Vital Básico", "BLS",
                "2024-01-10", "2026-01-10", "Cruz Roja", "Reentrenamiento anual"])
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=plantilla_cursos.xlsx"})


@talento_router.post("/importar-cursos")
async def importar_cursos(archivo: UploadFile = File(...),
                           db: Session = Depends(get_db),
                           current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    contenido = await archivo.read()
    ext = (archivo.filename or "").rsplit(".", 1)[-1].lower()
    filas = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        cabecera = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in fila):
                continue
            filas.append({cabecera[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(fila)})
    except Exception:
        if ext == "csv":
            texto = contenido.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(texto))
            filas = [{k.strip().lower(): v.strip() for k, v in row.items()} for row in reader]
        else:
            return RedirectResponse(url="/talento/vencimientos?error=excel", status_code=302)

    insertados, errores = 0, 0
    for f in filas:
        doc  = f.get("documento empleado (cédula)", f.get("documento", "")).strip()
        curso = f.get("nombre del curso", f.get("nombre_curso", "")).strip()
        vec_str = f.get("fecha vencimiento (aaaa-mm-dd)", f.get("fecha_vencimiento", "")).strip()
        if not doc or not curso or not vec_str:
            errores += 1
            continue
        emp = db.query(models.Empleado).filter_by(documento=doc, activo=True).first()
        if not emp:
            errores += 1
            continue
        fecha_vec  = _parse_fecha(vec_str)
        fecha_real = _parse_fecha(f.get("fecha realización (aaaa-mm-dd)", f.get("fecha_realizacion", "")))
        if not fecha_vec:
            errores += 1
            continue
        db.add(models.VencimientoCurso(
            empleado_id=emp.id, nombre_curso=curso,
            categoria=f.get("categoría", f.get("categoria", "")),
            fecha_realizacion=fecha_real, fecha_vencimiento=fecha_vec,
            institucion=f.get("institución", f.get("institucion", "")),
            observacion=f.get("observación", f.get("observacion", "")),
        ))
        insertados += 1
    db.commit()
    return RedirectResponse(url=f"/talento/vencimientos?importados={insertados}&errores={errores}", status_code=302)

@talento_router.get("/{emp_id}", response_class=HTMLResponse)
async def detalle_empleado(emp_id: int, request: Request, db: Session = Depends(get_db),
                            current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "talento", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    emp = db.query(models.Empleado).filter_by(id=emp_id).first()
    caps = db.query(models.Capacitacion).filter_by(empleado_id=emp_id).order_by(models.Capacitacion.fecha.desc()).all()
    vencimientos = db.query(models.VencimientoCurso).filter_by(empleado_id=emp_id, activo=True).order_by(models.VencimientoCurso.fecha_vencimiento).all()
    hoy = datetime.utcnow()
    return templates.TemplateResponse("talento/detalle.html", {
        "request": request, "user": current_user, "empleado": emp,
        "capacitaciones": caps, "vencimientos": vencimientos, "hoy": hoy,
        "permisos": auth.get_permisos(db, current_user),
    })

@talento_router.post("/{emp_id}/capacitacion")
async def agregar_capacitacion(
    emp_id: int, tema: str = Form(...), tipo: str = Form(""),
    fecha: str = Form(""), horas: float = Form(0),
    institucion: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    cap = models.Capacitacion(
        empleado_id=emp_id, tema=tema, tipo=tipo,
        horas=horas, institucion=institucion,
        fecha=datetime.fromisoformat(fecha) if fecha else None
    )
    db.add(cap)
    db.commit()
    return RedirectResponse(url=f"/talento/{emp_id}", status_code=302)

@talento_router.post("/{emp_id}/vencimiento")
async def agregar_vencimiento(
    emp_id: int,
    nombre_curso: str = Form(...), categoria: str = Form(""),
    fecha_realizacion: str = Form(""), fecha_vencimiento: str = Form(...),
    institucion: str = Form(""), observacion: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    if not current_user:
        return RedirectResponse(url="/login")
    vc = models.VencimientoCurso(
        empleado_id=emp_id,
        nombre_curso=nombre_curso, categoria=categoria,
        institucion=institucion, observacion=observacion,
        fecha_realizacion=datetime.fromisoformat(fecha_realizacion) if fecha_realizacion else None,
        fecha_vencimiento=datetime.fromisoformat(fecha_vencimiento),
    )
    db.add(vc)
    db.commit()
    return RedirectResponse(url=f"/talento/{emp_id}", status_code=302)

@talento_router.post("/{emp_id}/vencimiento/{vc_id}/eliminar")
async def eliminar_vencimiento(
    emp_id: int, vc_id: int,
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    if not current_user:
        return RedirectResponse(url="/login")
    vc = db.query(models.VencimientoCurso).filter_by(id=vc_id, empleado_id=emp_id).first()
    if vc:
        vc.activo = False
        db.commit()
    return RedirectResponse(url=f"/talento/{emp_id}", status_code=302)

# ─── PQRS ─────────────────────────────────────────────────────────────────────
pqrs_router = APIRouter(prefix="/pqrs", tags=["pqrs"])

@pqrs_router.get("", response_class=HTMLResponse)
async def lista_pqrs(request: Request, estado: str = "", tipo: str = "",
                      db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "pqrs", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    query = db.query(models.PQRS)
    if estado:
        query = query.filter(models.PQRS.estado == estado)
    if tipo:
        query = query.filter(models.PQRS.tipo == tipo)
    pqrs_list = query.order_by(models.PQRS.fecha_registro.desc()).all()
    return templates.TemplateResponse("pqrs/lista.html", {
        "request": request, "user": current_user,
        "pqrs_list": pqrs_list, "filtros": {"estado": estado, "tipo": tipo},
        "permisos": auth.get_permisos(db, current_user),
    })

@pqrs_router.get("/nueva", response_class=HTMLResponse)
async def nueva_pqrs_form(request: Request, db: Session = Depends(get_db),
                           current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "pqrs", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    usuarios = db.query(models.Usuario).filter_by(activo=True).all()
    return templates.TemplateResponse("pqrs/form.html", {
        "request": request, "user": current_user, "usuarios": usuarios, "pqrs": None,
        "permisos": auth.get_permisos(db, current_user),
    })

@pqrs_router.post("/nueva")
async def crear_pqrs(
    request: Request,
    codigo: str = Form(...), tipo: str = Form(...),
    nombre_solicitante: str = Form(""), email_solicitante: str = Form(""),
    telefono_solicitante: str = Form(""), servicio: str = Form(""),
    descripcion: str = Form(...), prioridad: str = Form("normal"),
    asignado_a: int = Form(None),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    from datetime import timedelta
    dias_limite = {"alta": 3, "normal": 10, "baja": 20}
    p = models.PQRS(
        codigo=codigo, tipo=tipo, nombre_solicitante=nombre_solicitante,
        email_solicitante=email_solicitante, telefono_solicitante=telefono_solicitante,
        servicio=servicio, descripcion=descripcion, prioridad=prioridad,
        asignado_a=asignado_a,
        fecha_limite=datetime.utcnow() + timedelta(days=dias_limite.get(prioridad, 10))
    )
    db.add(p)
    db.commit()
    return RedirectResponse(url="/pqrs", status_code=302)

@pqrs_router.post("/{pqrs_id}/responder")
async def responder_pqrs(
    pqrs_id: int, respuesta: str = Form(...), estado: str = Form("resuelta"),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    p = db.query(models.PQRS).filter_by(id=pqrs_id).first()
    if p:
        p.respuesta = respuesta
        p.estado = estado
        if estado in ["resuelta", "cerrada"]:
            p.fecha_cierre = datetime.utcnow()
        db.commit()
    return RedirectResponse(url="/pqrs", status_code=302)

# ─── FLUJOS ───────────────────────────────────────────────────────────────────
flujos_router = APIRouter(prefix="/flujos", tags=["flujos"])

@flujos_router.get("", response_class=HTMLResponse)
async def lista_flujos(request: Request, db: Session = Depends(get_db),
                        current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "flujos", "puede_ver"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    flujos = db.query(models.FlujoTrabajo).filter_by(activo=True).all()
    instancias = db.query(models.InstanciaFlujo).filter_by(estado="activo").order_by(
        models.InstanciaFlujo.fecha_inicio.desc()).limit(10).all()
    return templates.TemplateResponse("flujos/lista.html", {
        "request": request, "user": current_user,
        "flujos": flujos, "instancias": instancias,
        "permisos": auth.get_permisos(db, current_user),
    })

@flujos_router.get("/nuevo", response_class=HTMLResponse)
async def nuevo_flujo_form(request: Request, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login")
    if not auth.tiene_permiso(db, current_user, "flujos", "puede_crear"):
        return RedirectResponse(url="/?sin_acceso=1", status_code=302)
    return templates.TemplateResponse("flujos/form.html", {
        "request": request, "user": current_user, "flujo": None,
        "permisos": auth.get_permisos(db, current_user),
    })

@flujos_router.post("/nuevo")
async def crear_flujo(
    request: Request, nombre: str = Form(...),
    descripcion: str = Form(""), proceso: str = Form(""),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    f = models.FlujoTrabajo(nombre=nombre, descripcion=descripcion, proceso=proceso)
    db.add(f)
    db.commit()
    return RedirectResponse(url="/flujos", status_code=302)

@flujos_router.post("/{flujo_id}/iniciar")
async def iniciar_flujo(
    flujo_id: int, titulo: str = Form(...),
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    inst = models.InstanciaFlujo(
        flujo_id=flujo_id, titulo=titulo,
        iniciado_por=current_user.id
    )
    db.add(inst)
    db.commit()
    return RedirectResponse(url="/flujos", status_code=302)
